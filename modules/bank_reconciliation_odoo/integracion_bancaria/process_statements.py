import os
import datetime
import json
import csv
import re
import openpyxl
import io

# Google API imports (graceful fallback)
try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
    GOOGLE_DRIVE_AVAILABLE = True
except ImportError:
    GOOGLE_DRIVE_AVAILABLE = False

def parse_date(date_str):
    """
    Tries to parse date strings in various common formats and returns a YYYY-MM-DD string.
    """
    date_str = date_str.strip()
    for fmt in ('%d%m%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
        try:
            dt = datetime.datetime.strptime(date_str, fmt)
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            continue
    raise ValueError(f"Could not parse date string: {date_str}")

def parse_amount(amount_str):
    """
    Parses amount strings formatted with either dot/comma decimal/thousands separators.
    """
    amount_str = amount_str.strip()
    if not amount_str:
        return 0.0
    
    # Detect separator structure
    dot_idx = amount_str.rfind('.')
    comma_idx = amount_str.rfind(',')
    
    if comma_idx > dot_idx:
        # Comma is the decimal separator (Spanish locale)
        clean_str = amount_str.replace('.', '').replace(',', '.')
    else:
        # Dot is the decimal separator (English locale)
        clean_str = amount_str.replace(',', '')
        
    return float(clean_str)

def get_signed_amount(amount, token, movement_class, debit_tokens, credit_tokens):
    """
    Determines if the amount should be positive or negative based on the token
    or fallback movement class (NC=Credit/Positive, ND=Debit/Negative).
    """
    abs_amount = abs(amount)
    
    # 1. Specific bank token rules take priority
    if token in debit_tokens:
        return -abs_amount
    elif token in credit_tokens:
        return abs_amount
        
    # 2. Fallback to movement class (NC/ND)
    if movement_class == "ND":
        return -abs_amount
    elif movement_class == "NC":
        return abs_amount
        
    return amount

def clean_text(text):
    """
    Cleans label and reference strings of special/corrupt characters.
    Keeps alphanumeric characters, spaces, and simple punctuation (hyphen, dot, slash).
    """
    if not text:
        return ""
    # Remove special chars but keep spaces and common symbols
    cleaned = re.sub(r'[^a-zA-Z0-9\s\-\.\/]', '', text)
    # Remove duplicate spaces
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned.strip()

def parse_iso_datetime(iso_str):
    """
    Parses an ISO 8601 string from Google Drive API and returns a local date object.
    """
    if not iso_str:
        return None
    # Google API returns strings like "2026-06-11T20:58:35.000Z"
    # Replace Z with UTC offset
    iso_str = iso_str.replace('Z', '+00:00')
    try:
        dt = datetime.datetime.fromisoformat(iso_str)
        return dt.astimezone().date()
    except Exception:
        try:
            clean_str = iso_str.split('.')[0]
            dt = datetime.datetime.strptime(clean_str, "%Y-%m-%dT%H:%M:%S")
            return dt.date()
        except Exception:
            return None

def load_journal_mappings(excel_path):
    """
    Loads bank to journal mappings from the specified Excel file.
    Expects sheets containing 'VES' and 'USD' in their names.
    Returns a dictionary mapping (currency, bank_code) -> {'name': journal_name, 'id': journal_id}.
    """
    mappings = {}
    if not os.path.exists(excel_path):
        print(f"Warning: Excel mapping file not found at: {excel_path}")
        return mappings
        
    try:
        print(f"Loading journal mappings from Excel: {excel_path}")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        for sheet_name in wb.sheetnames:
            if "VES" in sheet_name:
                currency = "VES"
            elif "USD" in sheet_name:
                currency = "USD"
            else:
                currency = "VES" if "ves" in sheet_name.lower() else "USD"
                
            sheet = wb[sheet_name]
            header_row = [cell.value for cell in sheet[1]]
            
            def find_col_idx(names, default):
                for name in names:
                    for i, val in enumerate(header_row):
                        if val and name.lower() in str(val).lower():
                            return i
                return default
                
            code_idx = find_col_idx(["código", "codigo", "code"], 0)
            journal_idx = find_col_idx(["diario contable odoo", "diario", "journal"], 3)
            id_idx = find_col_idx(["id"], 4)
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= max(code_idx, journal_idx):
                    continue
                raw_code = row[code_idx]
                raw_journal = row[journal_idx]
                raw_id = row[id_idx] if len(row) > id_idx else None
                
                if raw_code is None:
                    continue
                    
                code = str(raw_code).strip()
                if code.isdigit():
                    code = code.zfill(4)
                    
                journal = str(raw_journal).strip() if raw_journal is not None else ""
                journal_id = int(raw_id) if raw_id is not None else None
                
                mappings[(currency, code)] = {
                    'name': journal,
                    'id': journal_id
                }
                
        print(f"Successfully loaded {len(mappings)} journal mappings.")
    except Exception as e:
        print(f"Error loading journal mappings: {e}")
        
    return mappings

def process_file(filepath, debit_tokens, credit_tokens, csv_headers, journal_mappings):
    """
    Parses the input TXT file, transforms transactions, looks up the Odoo journal,
    and writes the CSV output.
    """
    processed_rows = []
    
    print(f"Reading file: {filepath}")
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f, quotechar='"', delimiter=',')
        for line_num, row in enumerate(reader, 1):
            if not row:
                continue
                
            # Expecting 10 fields per row
            if len(row) < 10:
                print(f"Warning: Line {line_num} has fewer than 10 columns ({len(row)}). Skipping.")
                continue
                
            # Extract fields
            raw_bank_code = row[0].strip()
            raw_currency = row[1].strip()
            # row[2]: Bank Account
            raw_date = row[3]
            raw_ref = row[4]
            movement_class = row[5].strip()
            raw_label = row[6]
            raw_amount = row[7]
            # row[8]: Balance
            token = row[9].strip()
            
            # Skip Saldo Inicial (SI) and Saldo Final (SF) lines
            if movement_class in ("SI", "SF"):
                print(f"Line {line_num}: Skipping balance line ({movement_class}).")
                continue
                
            try:
                # Transform fields
                date_formatted = parse_date(raw_date)
                reference_clean = clean_text(raw_ref)
                label_clean = clean_text(raw_label)
                amount_parsed = parse_amount(raw_amount)
                
                amount_signed = get_signed_amount(
                    amount_parsed, token, movement_class, debit_tokens, credit_tokens
                )
                
                # Normalize code for mapping lookup
                bank_code = raw_bank_code
                if bank_code.isdigit():
                    bank_code = bank_code.zfill(4)
                
                # Normalize currency for mapping lookup (e.g. VEB/VEF -> VES)
                lookup_currency = raw_currency
                if lookup_currency in ("VEB", "VEF"):
                    lookup_currency = "VES"
                
                # Journal lookup
                journal_info = journal_mappings.get((lookup_currency, bank_code), {})
                journal_name = journal_info.get('name', '')
                journal_id = journal_info.get('id', None)
                
                if not journal_name:
                    print(f"Warning: No journal found in mapping for Currency={lookup_currency} (raw: {raw_currency}), Bank Code={bank_code}")
                
                processed_rows.append({
                    'Fecha': date_formatted,
                    'Referencia': reference_clean,
                    'Etiqueta': label_clean,
                    'Importe': f"{amount_signed:.2f}",
                    'Diario': journal_name,
                    'journal_id': journal_id
                })
                
                print(f"Line {line_num}: Processed {movement_class} reference {reference_clean} with amount {amount_signed:.2f} | Diario: {journal_name} (ID: {journal_id})")
                
            except Exception as e:
                print(f"Error processing line {line_num}: {e}. Skipping line.")
                
    return processed_rows

class GoogleDriveHelper:
    def __init__(self, creds_file, folder_id, creds_type='service_account'):
        self.creds_file = creds_file
        self.folder_id = folder_id
        self.creds_type = creds_type
        self.service = self._authenticate()
        
    def _authenticate(self):
        if not os.path.exists(self.creds_file):
            raise FileNotFoundError(f"Google credentials file not found at: {self.creds_file}")
        
        scopes = ['https://www.googleapis.com/auth/drive']
        
        if self.creds_type == 'oauth':
            token_path = os.path.join(os.path.dirname(self.creds_file), 'token.json')
            creds = None
            if os.path.exists(token_path):
                from google.oauth2.credentials import Credentials
                creds = Credentials.from_authorized_user_file(token_path, scopes)
            if not creds or not creds.valid:
                if creds and creds.expired and creds.refresh_token:
                    try:
                        from google.auth.transport.requests import Request
                        creds.refresh(Request())
                    except Exception:
                        creds = None
                if not creds:
                    from google_auth_oauthlib.flow import InstalledAppFlow
                    flow = InstalledAppFlow.from_client_secrets_file(self.creds_file, scopes)
                    creds = flow.run_local_server(port=0)
                with open(token_path, 'w') as token:
                    token.write(creds.to_json())
        else:
            # Default to service account
            creds = service_account.Credentials.from_service_account_file(
                self.creds_file,
                scopes=scopes
            )
            
        return build('drive', 'v3', credentials=creds)
        
    def list_files_in_folder(self, folder_id=None):
        fid = folder_id or self.folder_id
        query = f"'{fid}' in parents and trashed = false"
        results = self.service.files().list(
            q=query,
            fields="nextPageToken, files(id, name, mimeType, modifiedTime)",
            pageSize=100
        ).execute()
        return results.get('files', [])
        
    def download_file(self, file_id, local_path):
        print(f"Downloading file ID {file_id} to {local_path}...")
        request = self.service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
            
        with open(local_path, 'wb') as f:
            f.write(fh.getvalue())
            
    def upload_file(self, local_path, filename, folder_id=None, mimetype='text/csv'):
        fid = folder_id or self.folder_id
        query = f"'{fid}' in parents and name = '{filename}' and trashed = false"
        results = self.service.files().list(
            q=query,
            fields="files(id)",
            pageSize=1
        ).execute()
        existing_files = results.get('files', [])
        
        media = MediaFileUpload(local_path, mimetype=mimetype, resumable=True)
        
        if existing_files:
            file_id = existing_files[0]['id']
            print(f"Updating existing file '{filename}' (ID: {file_id}) in Google Drive folder {fid}...")
            self.service.files().update(
                fileId=file_id,
                media_body=media
            ).execute()
        else:
            print(f"Uploading new file '{filename}' to Google Drive folder {fid}...")
            file_metadata = {
                'name': filename,
                'parents': [fid]
            }
            self.service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()

def upload_to_odoo(rows, odoo_config, drive_helper=None, omitted_folder_id=None):
    """
    Uploads statement lines directly to Odoo via XML-RPC.
    Avoids duplicate entries by querying Odoo for existing records matching
    the same journal, date, reference, and amount.
    Writes omitted records to a local log file.
    """
    import xmlrpc.client
    
    url = odoo_config.get('url')
    db = odoo_config.get('db')
    user = odoo_config.get('user')
    password = odoo_config.get('password')
    
    if not all([url, db, user, password]):
        print("Error: Odoo configuration is incomplete in config.json. Skipping Odoo import.")
        return False
        
    print(f"\nConnecting to Odoo API at {url}...")
    try:
        common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
        uid = common.authenticate(db, user, password, {})
        if not uid:
            print("Error: Odoo authentication failed. Please check your config.json credentials.")
            return False
            
        print(f"Authenticated successfully with Odoo (UID: {uid}). Checking for duplicates...")
        models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")
        
        # 1. Collect all target journals and dates to query Odoo in a single batch
        target_journal_ids = set()
        target_dates = set()
        for row in rows:
            jid = row.get('journal_id')
            dt = row.get('Fecha')
            if jid:
                target_journal_ids.add(jid)
            if dt:
                target_dates.add(dt)
                
        # Fetch existing records in Odoo matching these journals and dates
        existing_set = set()
        if target_journal_ids and target_dates:
            print(f"Querying Odoo for existing statements in journals {list(target_journal_ids)} and dates {list(target_dates)}...")
            existing_records = models.execute_kw(
                db, uid, password,
                'account.bank.statement.line', 'search_read',
                [[
                    ('journal_id', 'in', list(target_journal_ids)),
                    ('date', 'in', list(target_dates))
                ]],
                {'fields': ['journal_id', 'date', 'ref', 'amount']}
            )
            for rec in existing_records:
                jid = rec['journal_id'][0] if rec['journal_id'] else None
                rdate = rec['date']
                rref = rec['ref'] or ''
                ramount = float(rec['amount']) if rec['amount'] is not None else 0.0
                
                # Use a rounded amount to avoid floating point precision mismatch
                existing_set.add((jid, rdate, rref, round(ramount, 2)))
                
        print(f"Found {len(existing_set)} existing record(s) in Odoo for matching date/journal windows.")
        
        # 2. Filter rows and separate into create vs omit lists
        records_to_create = []
        omitted_records = []
        
        for row in rows:
            journal_id = row.get('journal_id')
            if not journal_id:
                print(f"Skipping line reference {row.get('Referencia')}: No valid Journal ID mapped.")
                continue
                
            amount = float(row.get('Importe', 0.0))
            date = row.get('Fecha')
            ref = row.get('Referencia') or ''
            label = row.get('Etiqueta') or ''
            journal_name = row.get('Diario') or ''
            
            # Key for deduplication
            key = (journal_id, date, ref, round(amount, 2))
            
            if key in existing_set:
                omitted_records.append({
                    'Fecha': date,
                    'Referencia': ref,
                    'Etiqueta': label,
                    'Importe': f"{amount:.2f}",
                    'Diario': journal_name,
                    'journal_id': journal_id,
                    'Motivo': 'Ya existe en Odoo (Duplicado)'
                })
            else:
                vals = {
                    'journal_id': journal_id,
                    'date': date,
                    'payment_ref': label,
                    'ref': ref,
                    'amount': amount
                }
                records_to_create.append(vals)
                
        # 3. Log omitted records in a txt file
        base_dir = os.path.dirname(os.path.abspath(__file__))
        today_str = datetime.date.today().strftime("%Y-%m-%d")
        log_filename = f"registros_omitidos_{today_str}.txt"
        log_path = os.path.join(base_dir, log_filename)
        
        # Download existing log from Google Drive if configured to keep history
        if drive_helper and omitted_folder_id:
            try:
                print(f"Checking for existing {log_filename} in Google Drive folder {omitted_folder_id}...")
                query = f"'{omitted_folder_id}' in parents and name = '{log_filename}' and trashed = false"
                results = drive_helper.service.files().list(
                    q=query,
                    fields="files(id)",
                    pageSize=1
                ).execute()
                files = results.get('files', [])
                if files:
                    drive_file_id = files[0]['id']
                    print(f"Found existing log in Google Drive (ID: {drive_file_id}). Downloading to append...")
                    drive_helper.download_file(drive_file_id, log_path)
            except Exception as e:
                print(f"Warning: Could not check/download existing log file from Google Drive: {e}")
        
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(log_path, 'a', encoding='utf-8') as log_file:
            log_file.write(f"==================================================\n")
            log_file.write(f"Ejecucion: {now_str}\n")
            log_file.write(f"Total procesados para Odoo: {len(rows)}\n")
            log_file.write(f"Registros a crear: {len(records_to_create)}\n")
            log_file.write(f"Registros omitidos: {len(omitted_records)}\n")
            log_file.write(f"==================================================\n")
            if omitted_records:
                for idx, o_row in enumerate(omitted_records, 1):
                    log_file.write(
                        f"{idx}. Omitido: Diario: {o_row['Diario']} (ID: {o_row['journal_id']}) | "
                        f"Fecha: {o_row['Fecha']} | Ref: {o_row['Referencia']} | "
                        f"Etiqueta: {o_row['Etiqueta']} | Importe: {o_row['Importe']} | Motivo: {o_row['Motivo']}\n"
                    )
            else:
                log_file.write("No se omitieron registros en esta ejecucion.\n")
            log_file.write("\n")
            
        print(f"Omitted records log updated locally at: {log_path}")
        if omitted_records:
            print(f"Omitted {len(omitted_records)} duplicate record(s). Check '{log_filename}' for details.")
            
        # Upload updated log to Google Drive if configured
        if drive_helper and omitted_folder_id:
            try:
                print(f"Uploading updated {log_filename} to Google Drive folder {omitted_folder_id}...")
                drive_helper.upload_file(log_path, log_filename, folder_id=omitted_folder_id, mimetype='text/plain')
                if os.path.exists(log_path):
                    os.remove(log_path)
                    print(f"Deleted local {log_filename} after successful upload.")
            except Exception as e:
                print(f"Error uploading {log_filename} to Google Drive: {e}")
            
        # 4. Upload new records if any
        if not records_to_create:
            print("All records already exist in Odoo. No new entries created.")
            return True
            
        print(f"Uploading {len(records_to_create)} new statement line(s) to Odoo model 'account.bank.statement.line'...")
        new_ids = models.execute_kw(
            db, uid, password,
            'account.bank.statement.line', 'create',
            [records_to_create]
        )
        print(f"Successfully imported {len(new_ids)} new statement lines into Odoo! Generated IDs: {new_ids}")
        return True
        
    except Exception as e:
        print(f"An error occurred during Odoo direct import: {e}")
        return False

def main():
    # Load configuration
    base_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(base_dir, 'config.json')
    
    if not os.path.exists(config_path):
        print(f"Config file not found at {config_path}")
        exit(1)
        
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
        
    monitor_dir = config.get('monitor_dir', base_dir)
    test_mode = config.get('test_mode', False)
    debit_tokens = config.get('debit_tokens', [])
    credit_tokens = config.get('credit_tokens', [])
    csv_headers = config.get('csv_headers', ['Fecha', 'Referencia', 'Etiqueta', 'Importe', 'Diario'])
    
    # Target date determination
    filter_by_date = config.get('filter_by_date', True)
    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    
    match_dates = []
    date_strings = []
    
    if filter_by_date:
        if test_mode:
            match_dates = [today]
            print(f"--- RUNNING IN TEST MODE (Target: today {today}) ---")
        else:
            # Match files from today or yesterday to handle flexible upload times
            match_dates = [today, yesterday]
            print(f"--- RUNNING IN PRODUCTION MODE (Target: today {today} or yesterday {yesterday}) ---")
            
        # Generate all date strings to check in filenames
        for d in match_dates:
            date_strings.extend([
                d.strftime("%Y-%m-%d"),
                d.strftime("%Y%m%d"),
                d.strftime("%d%m%Y"),
                d.strftime("%d-%m-%Y")
            ])
    else:
        print("--- RUNNING WITHOUT DATE FILTER (Processing ALL TXT files) ---")
    
    # Load journal mappings
    excel_path = os.path.join(base_dir, 'Bancos_Venezolanos VES y USD.xlsx')
    journal_mappings = load_journal_mappings(excel_path)
    
    # Google Drive configuration
    use_google_drive = config.get('use_google_drive', False)
    google_drive_folder_id = config.get('google_drive_folder_id')
    google_creds_file = config.get('google_creds_file', 'service_account.json')
    google_creds_type = config.get('google_creds_type', 'service_account')
    
    drive_helper = None
    if use_google_drive:
        if not GOOGLE_DRIVE_AVAILABLE:
            print("Warning: Google API libraries are not installed. Google Drive integration disabled.")
            use_google_drive = False
        elif not google_drive_folder_id:
            print("Warning: google_drive_folder_id is not specified in config.json. Google Drive integration disabled.")
            use_google_drive = False
        else:
            try:
                creds_path = os.path.join(base_dir, google_creds_file)
                drive_helper = GoogleDriveHelper(creds_path, google_drive_folder_id, google_creds_type)
                print(f"Successfully authenticated with Google Drive API ({google_creds_type}).")
            except Exception as e:
                print(f"Error authenticating with Google Drive API: {e}")
                print("Falling back to local directory monitoring.")
                use_google_drive = False
                
    # Prepare directories
    temp_dir = os.path.join(base_dir, 'temp_drive')
    if use_google_drive:
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
            
    candidate_files = [] # list of tuples: (local_filepath, drive_filename_or_none)
    
    try:
        if use_google_drive:
            print("Scanning Google Drive folder for files...")
            try:
                drive_files = drive_helper.list_files_in_folder()
                print(f"Found {len(drive_files)} file(s) total in Google Drive folder.")
                
                for file in drive_files:
                    filename = file['name']
                    if not filename.endswith('.txt'):
                        continue
                        
                    if filter_by_date:
                        name_matches = any(ds in filename for ds in date_strings)
                        mtime_date = parse_iso_datetime(file.get('modifiedTime'))
                        mtime_matches = (mtime_date in match_dates)
                        should_process = name_matches or mtime_matches
                    else:
                        should_process = True
                    
                    if should_process:
                        local_path = os.path.join(temp_dir, filename)
                        drive_helper.download_file(file['id'], local_path)
                        candidate_files.append((local_path, filename))
            except Exception as e:
                print(f"Error scanning Google Drive files: {e}")
        else:
            # Scan local directory
            if not os.path.exists(monitor_dir):
                print(f"Monitor directory {monitor_dir} does not exist. Creating it.")
                os.makedirs(monitor_dir)
                
            print(f"Scanning local monitor directory: {monitor_dir}")
            for filename in os.listdir(monitor_dir):
                if not filename.endswith('.txt'):
                    continue
                filepath = os.path.join(monitor_dir, filename)
                if not os.path.isfile(filepath):
                    continue
                    
                if filter_by_date:
                    name_matches = any(ds in filename for ds in date_strings)
                    mtime = os.path.getmtime(filepath)
                    mtime_date = datetime.date.fromtimestamp(mtime)
                    mtime_matches = (mtime_date in match_dates)
                    should_process = name_matches or mtime_matches
                else:
                    should_process = True
                    
                if should_process:
                    candidate_files.append((filepath, None))
                    
        if not candidate_files:
            if filter_by_date:
                print(f"No files matched target dates: {', '.join(str(d) for d in match_dates)}. Silently exiting.")
            else:
                print("No files found to process. Silently exiting.")
            return
            
        print(f"Found {len(candidate_files)} file(s) to process.")
        
        all_processed_rows = []
        
        for filepath, drive_filename in candidate_files:
            rows = process_file(filepath, debit_tokens, credit_tokens, csv_headers, journal_mappings)
            
            if not rows:
                print(f"No transactable lines processed for {filepath}.")
                continue
                
            all_processed_rows.extend(rows)
            
            # Write CSV output
            base_name, _ = os.path.splitext(filepath)
            csv_path = base_name + '.csv'
            
            print(f"Writing output CSV to: {csv_path}")
            with open(csv_path, 'w', newline='', encoding='utf-8') as csv_file:
                writer = csv.DictWriter(csv_file, fieldnames=csv_headers, extrasaction='ignore')
                writer.writeheader()
                for row in rows:
                    writer.writerow(row)
                    
            print(f"Successfully generated template CSV locally for {os.path.basename(filepath)}")
            
            if use_google_drive and drive_filename:
                csv_filename = os.path.splitext(drive_filename)[0] + '.csv'
                try:
                    drive_helper.upload_file(csv_path, csv_filename)
                    print(f"Successfully uploaded CSV to Google Drive: {csv_filename}")
                except Exception as e:
                    print(f"Error uploading {csv_filename} to Google Drive: {e}")
                    
        # Direct Odoo import
        use_odoo_import = config.get('use_odoo_import', False)
        if use_odoo_import and all_processed_rows:
            print("\nStarting direct Odoo import...")
            omitted_folder_id = config.get('google_drive_omitted_folder_id') if use_google_drive else None
            upload_to_odoo(
                all_processed_rows,
                config.get('odoo', {}),
                drive_helper=drive_helper,
                omitted_folder_id=omitted_folder_id
            )
                    
    finally:
        # Clean up any processed candidate files and their generated CSVs to keep local folders clean
        if candidate_files:
            print("Cleaning up local processed files...")
            for filepath, _ in candidate_files:
                # Delete the source TXT file
                if os.path.exists(filepath):
                    try:
                        os.remove(filepath)
                        print(f"Deleted local file: {filepath}")
                    except Exception as e:
                        print(f"Error deleting local file {filepath}: {e}")
                
                # Delete the generated CSV file
                base_name, _ = os.path.splitext(filepath)
                csv_path = base_name + '.csv'
                if os.path.exists(csv_path):
                    try:
                        os.remove(csv_path)
                        print(f"Deleted local CSV: {csv_path}")
                    except Exception as e:
                        print(f"Error deleting local CSV {csv_path}: {e}")

        # Clean up temporary downloads directory if it exists
        if os.path.exists(temp_dir):
            print("Cleaning up temporary directory...")
            for filename in os.listdir(temp_dir):
                file_path = os.path.join(temp_dir, filename)
                try:
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                except Exception as e:
                    print(f"Error deleting temp file {file_path}: {e}")
            try:
                os.rmdir(temp_dir)
            except Exception as e:
                print(f"Error removing temp directory {temp_dir}: {e}")

if __name__ == '__main__':
    main()

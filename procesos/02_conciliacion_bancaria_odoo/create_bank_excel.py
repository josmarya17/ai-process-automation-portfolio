import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel():
    # Data list of Venezuelan Banks and their 4-digit codes
    banks_data = [
        {"code": "0001", "name": "Banco Central de Venezuela (BCV)", "type": "Banco Central", "status": "Activo"},
        {"code": "0102", "name": "Banco de Venezuela", "type": "Banco Universal (Público)", "status": "Activo"},
        {"code": "0104", "name": "Banco Venezolano de Crédito", "type": "Banco Universal", "status": "Activo"},
        {"code": "0105", "name": "Banco Mercantil", "type": "Banco Universal", "status": "Activo"},
        {"code": "0108", "name": "BBVA Provincial", "type": "Banco Universal", "status": "Activo"},
        {"code": "0114", "name": "Bancaribe", "type": "Banco Universal", "status": "Activo"},
        {"code": "0115", "name": "Banco Exterior", "type": "Banco Universal", "status": "Activo"},
        {"code": "0128", "name": "Banco Caroní", "type": "Banco Universal", "status": "Activo"},
        {"code": "0134", "name": "Banesco", "type": "Banco Universal", "status": "Activo"},
        {"code": "0137", "name": "Banco Sofitasa", "type": "Banco Universal", "status": "Activo"},
        {"code": "0138", "name": "Banco Plaza", "type": "Banco Universal", "status": "Activo"},
        {"code": "0146", "name": "Bangente", "type": "Banco Microfinanciero", "status": "Activo"},
        {"code": "0151", "name": "Banco Fondo Común (BFC)", "type": "Banco Universal", "status": "Activo"},
        {"code": "0156", "name": "100% Banco", "type": "Banco Universal", "status": "Activo"},
        {"code": "0157", "name": "DelSur Banco Universal", "type": "Banco Universal", "status": "Activo"},
        {"code": "0163", "name": "Banco del Tesoro", "type": "Banco Universal (Público)", "status": "Activo"},
        {"code": "0166", "name": "Banco Agrícola de Venezuela", "type": "Banco Universal (Público)", "status": "Activo"},
        {"code": "0168", "name": "Bancrecer", "type": "Banco Microfinanciero", "status": "Activo"},
        {"code": "0169", "name": "Mi Banco", "type": "Banco Microfinanciero", "status": "Activo"},
        {"code": "0171", "name": "Banco Activo", "type": "Banco Universal", "status": "Activo"},
        {"code": "0172", "name": "Bancamiga", "type": "Banco Universal", "status": "Activo"},
        {"code": "0173", "name": "Banco Internacional de Desarrollo (BID)", "type": "Banco Universal", "status": "Activo"},
        {"code": "0174", "name": "Banplus", "type": "Banco Universal", "status": "Activo"},
        {"code": "0175", "name": "Banco Digital de los Trabajadores (BDT)", "type": "Banco Universal (Público)", "status": "Activo"},
        {"code": "0177", "name": "BANFANB", "type": "Banco Universal (Público)", "status": "Activo"},
        {"code": "0191", "name": "Banco Nacional de Crédito (BNC)", "type": "Banco Universal", "status": "Activo"},
        {"code": "0601", "name": "Instituto Municipal de Crédito Popular (IMCP)", "type": "Entidad Financiera Municipal", "status": "Activo"}
    ]
    
    # Initialize workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bancos Venezolanos"
    
    # Enable gridlines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Header titles
    headers = ["Código", "Banco / Institución", "Tipo de Entidad", "Estado / Estatus"]
    
    # Style definitions
    font_name = "Segoe UI"
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=10, bold=False, color="333333")
    
    # Premium Blue/Teal Palette Fills
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Alignments
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Borders
    thin_side = Side(border_style="thin", color="D3D3D3")
    double_bottom_side = Side(border_style="double", color="1F4E78")
    thin_top_side = Side(border_style="thin", color="1F4E78")
    
    data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=double_bottom_side)
    
    # Write Headers
    ws.append(headers)
    for col_num in range(1, 5):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = header_border
        
    # Write Data
    for row_idx, data in enumerate(banks_data, 2):
        row_values = [data["code"], data["name"], data["type"], data["status"]]
        ws.append(row_values)
        
        # Style row cells
        is_even = (row_idx % 2 == 0)
        current_fill = zebra_fill if is_even else white_fill
        
        for col_num in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = data_font
            cell.fill = current_fill
            cell.border = data_border
            
            # Alignments & formatting
            if col_num == 1:
                # Format as TEXT so Excel doesn't drop leading zeros
                cell.number_format = "@"
                cell.alignment = center_align
            elif col_num == 4:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # Add a thin total/summary border at the bottom
    for col_num in range(1, 5):
        cell = ws.cell(row=len(banks_data) + 2, column=col_num)
        # This will be an empty cell under the table, just to close it visually
        cell.border = Border(top=thin_top_side)
        
    # Set Row Heights
    ws.row_dimensions[1].height = 28
    for r in range(2, len(banks_data) + 2):
        ws.row_dimensions[r].height = 20
        
    # Auto-fit Column Widths with padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        # Add a padding of 5 characters
        ws.column_dimensions[col_letter].width = max(max_len + 5, 12)
        
    # Save Workbook
    filename = "Bancos_Venezolanos.xlsx"
    wb.save(filename)
    print(f"Successfully generated styled Excel: {filename}")

if __name__ == "__main__":
    create_excel()
